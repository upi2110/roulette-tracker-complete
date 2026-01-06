"""
EXCEL EXPORT UTILITY
Creates comprehensive Excel reports from backtest results
with Session Results, Analytics Dashboard, and Detailed Trades sheets
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from typing import List, Dict
from datetime import datetime


class BacktestExporter:
    """Export backtest results to formatted Excel file"""
    
    def __init__(self):
        self.wb = Workbook()
        
        # Define colors
        self.color_header = "366092"  # Dark blue
        self.color_success = "C6EFCE"  # Light green
        self.color_fail = "FFC7CE"     # Light red
        self.color_warning = "FFEB9C"  # Light yellow
        
    def export_results(self, results: List[Dict], analytics: Dict, filename: str):
        """
        Export complete backtest results to Excel
        
        Args:
            results: List of session results from run_full_backtest()
            analytics: Analytics dict from calculate_analytics()
            filename: Output filename
        """
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        
        # Create sheets
        self._create_session_results_sheet(results)
        self._create_analytics_dashboard_sheet(analytics, results)
        self._create_detailed_trades_sheet(results)
        
        # Save workbook
        self.wb.save(filename)
        return filename
    
    def _create_session_results_sheet(self, results: List[Dict]):
        """Create Session Results sheet with summary of each test"""
        ws = self.wb.create_sheet("Session Results", 0)
        
        # Headers
        headers = [
            "Test #", "Start Pos", "Start Number", "Result", 
            "Spins Needed", "Final Bankroll", "Profit/Loss",
            "Total Bets", "Wins", "Losses", "Win Rate %",
            "Max Drawdown", "Peak Bankroll"
        ]
        
        ws.append(headers)
        
        # Format header row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(1, col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.color_header, end_color=self.color_header, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data rows
        for result in results:
            profit_loss = result['final_bankroll'] - 4000  # Assuming $4000 starting
            
            row_data = [
                result['test_number'],
                result['start_position'],
                result['start_number'] if result['start_number'] is not None else "N/A",
                result['result'],
                result['spins_needed'],
                f"${result['final_bankroll']:,.0f}",
                f"${profit_loss:+,.0f}",
                result['total_bets'],
                result['wins'],
                result['losses'],
                f"{result['win_rate']:.1f}%",
                f"${result['max_drawdown']:,.0f}",
                f"${result['peak_bankroll']:,.0f}"
            ]
            
            ws.append(row_data)
            
            # Color code result column
            row_num = ws.max_row
            result_cell = ws.cell(row_num, 4)
            
            if result['result'] == 'TARGET_REACHED':
                result_cell.fill = PatternFill(start_color=self.color_success, end_color=self.color_success, fill_type="solid")
            elif result['result'] in ['BANKRUPT', 'MAX_SPINS_EXCEEDED']:
                result_cell.fill = PatternFill(start_color=self.color_fail, end_color=self.color_fail, fill_type="solid")
            else:
                result_cell.fill = PatternFill(start_color=self.color_warning, end_color=self.color_warning, fill_type="solid")
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"
    
    def _create_analytics_dashboard_sheet(self, analytics: Dict, results: List[Dict]):
        """Create Analytics Dashboard with summary statistics"""
        ws = self.wb.create_sheet("Analytics Dashboard", 1)
        
        # Title
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = "BACKTEST ANALYTICS DASHBOARD"
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color=self.color_header, end_color=self.color_header, fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.row_dimensions[1].height = 30
        
        # Overview section
        row = 3
        ws.cell(row, 1, "OVERVIEW").font = Font(bold=True, size=12)
        row += 1
        
        overview_data = [
            ("Total Sessions Tested:", analytics['total_sessions']),
            ("Successful Sessions:", f"{analytics['successful_count']} ({analytics['success_rate']:.1f}%)"),
            ("Failed (Bankrupt):", f"{analytics['failed_bankrupt_count']}"),
            ("Failed (Max Spins):", f"{analytics['failed_max_spins_count']}"),
            ("Overall Failure Rate:", f"{analytics['failure_rate']:.1f}%"),
        ]
        
        for label, value in overview_data:
            ws.cell(row, 1, label).font = Font(bold=True)
            ws.cell(row, 2, value)
            row += 1
        
        # Performance Metrics
        row += 2
        ws.cell(row, 1, "PERFORMANCE METRICS").font = Font(bold=True, size=12)
        row += 1
        
        performance_data = [
            ("Average Spins to Win:", f"{analytics['avg_spins_to_win']:.1f}"),
            ("Minimum Spins to Win:", analytics['min_spins_to_win']),
            ("Maximum Spins to Win:", analytics['max_spins_to_win']),
            ("Average Spins to Lose:", f"{analytics['avg_spins_to_lose']:.1f}"),
            ("Average Win Rate (Successful):", f"{analytics['avg_win_rate']:.1f}%"),
            ("Overall Win Rate (All Trades):", f"{analytics['overall_win_rate']:.1f}%"),
        ]
        
        for label, value in performance_data:
            ws.cell(row, 1, label).font = Font(bold=True)
            ws.cell(row, 2, value)
            row += 1
        
        # Financial Metrics
        row += 2
        ws.cell(row, 1, "FINANCIAL METRICS").font = Font(bold=True, size=12)
        row += 1
        
        financial_data = [
            ("Average Final Bankroll:", f"${analytics['avg_final_bankroll']:,.0f}"),
            ("Average Max Drawdown:", f"${analytics['avg_max_drawdown']:,.0f}"),
            ("Maximum Drawdown (All Sessions):", f"${analytics['max_drawdown_overall']:,.0f}"),
            ("Total Trades Made:", f"{analytics['total_trades']:,}"),
            ("Total Wins:", f"{analytics['total_wins']:,}"),
            ("Total Losses:", f"{analytics['total_losses']:,}"),
        ]
        
        for label, value in financial_data:
            ws.cell(row, 1, label).font = Font(bold=True)
            ws.cell(row, 2, value)
            row += 1
        
        # Best and Worst Sessions
        if analytics['best_session'] and analytics['worst_session']:
            row += 2
            ws.cell(row, 1, "BEST & WORST SESSIONS").font = Font(bold=True, size=12)
            row += 1
            
            best = analytics['best_session']
            worst = analytics['worst_session']
            
            best_worst_data = [
                ("Best Session:", f"Test #{best['test_number']} - ${best['final_bankroll']:,.0f} in {best['spins_needed']} spins"),
                ("Worst Session:", f"Test #{worst['test_number']} - ${worst['final_bankroll']:,.0f} in {worst['spins_needed']} spins"),
            ]
            
            for label, value in best_worst_data:
                ws.cell(row, 1, label).font = Font(bold=True)
                ws.cell(row, 2, value)
                row += 1
        
        # Distribution Analysis
        row += 2
        ws.cell(row, 1, "DISTRIBUTION ANALYSIS").font = Font(bold=True, size=12)
        row += 1
        
        # Count sessions by spin ranges
        spin_ranges = {
            '0-25': 0,
            '26-50': 0,
            '51-100': 0,
            '101-150': 0,
            '151-200': 0,
            '201-250': 0
        }
        
        for result in results:
            spins = result['spins_needed']
            if spins <= 25:
                spin_ranges['0-25'] += 1
            elif spins <= 50:
                spin_ranges['26-50'] += 1
            elif spins <= 100:
                spin_ranges['51-100'] += 1
            elif spins <= 150:
                spin_ranges['101-150'] += 1
            elif spins <= 200:
                spin_ranges['151-200'] += 1
            else:
                spin_ranges['201-250'] += 1
        
        ws.cell(row, 1, "Spins Range").font = Font(bold=True)
        ws.cell(row, 2, "Session Count").font = Font(bold=True)
        row += 1
        
        for spin_range, count in spin_ranges.items():
            ws.cell(row, 1, spin_range)
            ws.cell(row, 2, count)
            row += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 35
    
    def _create_detailed_trades_sheet(self, results: List[Dict]):
        """Create Detailed Trades sheet with all individual bets"""
        ws = self.wb.create_sheet("Detailed Trades", 2)
        
        # Headers
        headers = [
            "Session #", "Spin #", "Number Hit", "Predicted Numbers",
            "Bet/Number", "Total Bet", "Hit/Miss", "Profit/Loss",
            "Bankroll", "Confidence %"
        ]
        
        ws.append(headers)
        
        # Format header row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(1, col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.color_header, end_color=self.color_header, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add all trades from all sessions
        for result in results:
            session_num = result['test_number']
            
            for trade in result['trades']:
                # Format prediction as comma-separated
                prediction_str = ', '.join(str(n) for n in trade['prediction'][:5])
                if len(trade['prediction']) > 5:
                    prediction_str += f"... ({len(trade['prediction'])} total)"
                
                row_data = [
                    session_num,
                    trade['spin'],
                    trade['number_hit'],
                    prediction_str,
                    f"${trade['bet_amount']:.0f}",
                    f"${trade['total_bet']:.0f}",
                    "HIT" if trade['hit'] else "MISS",
                    f"${trade['profit']:+,.0f}",
                    f"${trade['bankroll']:,.0f}",
                    f"{trade['confidence']*100:.1f}%"
                ]
                
                ws.append(row_data)
                
                # Color code hit/miss
                row_num = ws.max_row
                hit_cell = ws.cell(row_num, 7)
                profit_cell = ws.cell(row_num, 8)
                
                if trade['hit']:
                    hit_cell.fill = PatternFill(start_color=self.color_success, end_color=self.color_success, fill_type="solid")
                    profit_cell.fill = PatternFill(start_color=self.color_success, end_color=self.color_success, fill_type="solid")
                else:
                    hit_cell.fill = PatternFill(start_color=self.color_fail, end_color=self.color_fail, fill_type="solid")
                    profit_cell.fill = PatternFill(start_color=self.color_fail, end_color=self.color_fail, fill_type="solid")
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Add filter to headers
        ws.auto_filter.ref = ws.dimensions
